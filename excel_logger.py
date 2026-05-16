"""
excel_logger.py
===============
Drop this file next to bot.py on your server.

It appends a new row to CS_GO_PnL_Tracker_Final.xlsx whenever an item sells.
Fills in: Item Name, Cost (USD), Sold Price (USD)
Leaves blank for you to fill: Link
All formula columns (Fee, Net Sold, Profit, ROI%) are pre-formatted in the sheet.

Usage (called automatically by the bot):
    from excel_logger import log_sale
    log_sale("MP9 | Sand Dashed 0.130889981985", sold_price_cents=410, cost_cents=310)
"""

import logging
import os
from pathlib import Path
from openpyxl import load_workbook

# ── Config ─────────────────────────────────────────────────────────────────────
EXCEL_FILE  = os.getenv("EXCEL_PATH", "CS_GO_PnL_Tracker_Final.xlsx")
SHEET_NAME  = "CSGO PnL Tracker"

# Column positions (1-indexed), matching your sheet:
# A=Item Name, B=Cost, C=Sold Price, D=Fee Amount, E=Net Sold, F=Profit, G=ROI%, H=Link
COL_NAME = 1  # A
COL_COST = 2  # B
COL_SOLD = 3  # C
COL_FEE  = 4  # D

# Your sheet's summary block starts at row 1149, so data rows are 3-1147
DATA_START_ROW = 3
DATA_END_ROW   = 1145  # hard cap - never write into or past the summary
log = logging.getLogger("csfloat_bot.excel_logger")


def _find_next_row(sheet) -> int | None:
    """Return the first empty row within the data range (rows 3-1145)."""
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        if sheet.cell(row=row, column=COL_NAME).value is None:
            return row
    log.warning("Data range is full (row %s). Expand your sheet.", DATA_END_ROW)
    return None


def log_sale(item_name: str, sold_price_cents: int, cost_cents: int = None) -> bool:
    """
    Append a sold item row to the Excel tracker.

    Args:
        item_name:         The display name from config.json
        sold_price_cents:  The reserve_price (in cents, e.g. 410 = $4.10)
        cost_cents:        The cost field from config.json (optional, in cents)

    Returns:
        True on success, False on error.
    """
    excel_path = Path(EXCEL_FILE)
    if not excel_path.exists():
        log.error("'%s' not found - skipping log.", EXCEL_FILE)
        return False

    try:
        sold_price_usd = sold_price_cents / 100.0
        cost_usd       = (cost_cents / 100.0) if cost_cents else None

        wb = load_workbook(excel_path)

        if SHEET_NAME not in wb.sheetnames:
            log.error("Sheet '%s' not found.", SHEET_NAME)
            return False

        ws = wb[SHEET_NAME]

        # Find the next empty row in the data range
        r = _find_next_row(ws)
        if r is None:
            return False

        # The sheet has pre-formatted rows with formulas already in D-G.
        # We only need to write A (name), B (cost), C (sold price).
        # The existing formulas handle fee, net, profit, ROI automatically.
        ws.cell(row=r, column=COL_NAME).value = item_name
        ws.cell(row=r, column=COL_COST).value = cost_usd
        ws.cell(row=r, column=COL_SOLD).value = sold_price_usd

        # Only add formulas if the row is not already pre-formatted
        if ws.cell(row=r, column=COL_FEE).value is None:
            fee_ref = "$B$1"
            ws.cell(row=r, column=4).value = f"=IF(C{r}=\"\",\"\",C{r}*{fee_ref})"
            ws.cell(row=r, column=5).value = f"=IF(C{r}=\"\",\"\",C{r}-D{r})"
            ws.cell(row=r, column=6).value = f"=IF(OR(B{r}=\"\",C{r}=\"\"),\"\",E{r}-B{r})"
            ws.cell(row=r, column=7).value = f"=IF(OR(B{r}=\"\",B{r}=0),\"\",F{r}/B{r})"

        wb.save(excel_path)
        log.info("Logged sale -> row %s: '%s' @ $%.2f cost=$%s", r, item_name, sold_price_usd, cost_usd)
        return True

    except Exception as e:
        log.error("Error writing to Excel: %s", e)
        return False
